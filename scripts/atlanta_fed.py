"""Atlanta Fed Wage Growth Tracker: job switcher vs stayer premium.

The tracker publishes median 12-month wage growth for job switchers and job
stayers (3-month moving averages) as a downloadable Excel workbook. The
switcher premium — switcher minus stayer, in percentage points — is the
dashboard's "does moving actually pay right now?" component.

This fetcher is deliberately paranoid: the workbook's layout is not a stable
API, so data is located by header text, ambiguity is treated as failure, and
parsed values must pass magnitude sanity checks. Every failure raises rather
than guesses; callers fall back to last-known-good values (update_data.py),
so a broken download or a reshuffled spreadsheet can never blank out or
corrupt the published series. verify_sources.py dumps the workbook structure
so a human can adjust the matching rules precisely when the layout changes.

Requires openpyxl (installed by the GitHub Actions workflow); import is
deferred so the FRED-only path has no third-party dependency.
"""

import io
import re
from datetime import datetime
from urllib.parse import urljoin
from urllib.request import Request, urlopen

# Direct workbook guesses, tried first.
WORKBOOK_URLS = [
    'https://www.atlantafed.org/-/media/documents/datafiles/chcs/wage-growth-tracker/wage-growth-data.xlsx',
]
# The tracker's landing page, scraped for the current workbook link when the
# direct guesses fail or stop pointing at a real xlsx.
DISCOVERY_PAGES = [
    'https://www.atlantafed.org/chcs/wage-growth-tracker',
]
_UA = {'User-Agent': 'Mozilla/5.0 (compatible; labor-market-dashboard/1.0; '
                     '+https://github.com/jricciardi/labor-market-dashboard)'}

_SWITCHER = re.compile(r'switcher', re.I)
_STAYER = re.compile(r'stayer', re.I)
_XLSX_HREF = re.compile(r'href="([^"]*wage[-_]?growth[^"]*\.xlsx?)"', re.I)
_ZIP_MAGIC = b'PK\x03\x04'

# Sanity bounds: the tracker's median wage growth series has lived in roughly
# 0-16% for its entire history; the premium in roughly -3..+4pp. Values far
# outside mean we parsed the wrong thing (e.g. Excel percent fractions).
_WAGE_RANGE = (-5.0, 25.0)
_PREMIUM_RANGE = (-6.0, 8.0)
_MIN_PLAUSIBLE_MEDIAN = 0.5  # fractions like 0.045 would fail this


class AtlantaFedError(RuntimeError):
    """Raised when the tracker workbook can't be fetched or understood."""


def _download(url, timeout=60):
    with urlopen(Request(url, headers=_UA), timeout=timeout) as resp:
        return resp.read()


def _workbook_bytes():
    """Fetch the workbook, discovering the current link if guesses fail.

    Returns (content, url). Raises AtlantaFedError with per-URL diagnostics.
    """
    errors = []
    candidates = list(WORKBOOK_URLS)
    for page_url in DISCOVERY_PAGES:
        try:
            page = _download(page_url).decode('utf-8', 'replace')
            for href in _XLSX_HREF.findall(page):
                candidates.append(urljoin(page_url, href))
        except Exception as e:  # noqa: BLE001 - discovery is best-effort
            errors.append(f'discovery {page_url}: {e}')
    seen = set()
    for url in candidates:
        if url in seen:
            continue
        seen.add(url)
        try:
            content = _download(url)
        except Exception as e:  # noqa: BLE001
            errors.append(f'{url}: {e}')
            continue
        if not content.startswith(_ZIP_MAGIC):
            head = content[:40].decode('utf-8', 'replace')
            errors.append(f'{url}: not an xlsx (starts with {head!r})')
            continue
        return content, url
    raise AtlantaFedError('no workbook found; ' + '; '.join(errors))


def _parse_date(cell):
    """Tracker date cells: datetimes, 'YYYY-MM', or 'Mon-YY' strings."""
    if isinstance(cell, datetime):
        return cell.strftime('%Y-%m-01')
    if isinstance(cell, str):
        for fmt in ('%Y-%m-%d', '%Y-%m', '%b-%y', '%b-%Y', '%m/%d/%Y'):
            try:
                return datetime.strptime(cell.strip(), fmt).strftime('%Y-%m-01')
            except ValueError:
                continue
    return None


def _find_columns(rows):
    """Locate (header_row_idx, switcher_col, stayer_col) by header text.

    Requires exactly one column matching /switcher/i and exactly one matching
    /stayer/i in the same header row — ambiguity (e.g. smoothed and
    unsmoothed variants side by side, or a derived "switcher minus stayer"
    column) raises instead of guessing a column.
    """
    for row_idx, row in enumerate(rows[:10]):
        switchers, stayers = [], []
        for col_idx, cell in enumerate(row):
            if not isinstance(cell, str):
                continue
            if _SWITCHER.search(cell):
                switchers.append(col_idx)
            elif _STAYER.search(cell):
                stayers.append(col_idx)
        if switchers or stayers:
            if len(switchers) == 1 and len(stayers) == 1:
                return row_idx, switchers[0], stayers[0]
            raise AtlantaFedError(
                f'ambiguous switcher/stayer headers in row {row_idx}: '
                f'{len(switchers)} switcher and {len(stayers)} stayer columns')
    return None


def _sanity_check(series):
    """Reject parses whose magnitudes can't be median wage growth percents."""
    values = [v for pair in series.values() for v in pair]
    values.sort()
    median = values[len(values) // 2]
    if not (_MIN_PLAUSIBLE_MEDIAN <= median <= _WAGE_RANGE[1]):
        raise AtlantaFedError(
            f'wage growth magnitudes implausible (median {median}); '
            'wrong column or percent-fraction formatting?')
    if any(not (_WAGE_RANGE[0] <= v <= _WAGE_RANGE[1]) for v in values):
        raise AtlantaFedError('wage growth values outside plausible range')
    premiums = [sw - st for sw, st in series.values()]
    if any(not (_PREMIUM_RANGE[0] <= p <= _PREMIUM_RANGE[1]) for p in premiums):
        raise AtlantaFedError('switcher premium outside plausible range')


# Sheet preference, most-preferred first (verified against the live workbook
# structure, 2026-07-16): `data_overall` is the tracker's headline 3MMA grid
# and carries Job Switcher / Job Stayer columns; the `Job Switcher` sheet is
# the 12-month-average cut (small-sample smoothing, laggier); anything else
# that happens to match (`Alternative WGT` uses a different methodology) is
# never used.
SHEET_PREFERENCE = [re.compile(r'^data_overall$', re.I),
                    re.compile(r'^job switcher$', re.I)]


def parse_workbook(content):
    """Extract {YYYY-MM-01: (switcher, stayer)} from workbook bytes.

    Only sheets on the SHEET_PREFERENCE list are eligible, in order — a
    workbook where none of the preferred sheets parses is an error (never
    fall through to an arbitrary sheet: it could be an alternative
    methodology or an unsmoothed cut).
    """
    import openpyxl  # deferred: only the switcher path needs it

    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    by_title = {s.title: s for s in workbook.worksheets}
    notes = []
    for pattern in SHEET_PREFERENCE:
        for title, sheet in by_title.items():
            if not pattern.match(title):
                continue
            rows = [list(r) for r in sheet.iter_rows(values_only=True)]
            try:
                found = _find_columns(rows)
            except AtlantaFedError as e:
                notes.append(f'{title}: {e}')
                continue
            if not found:
                notes.append(f'{title}: no switcher/stayer headers')
                continue
            header_idx, sw_col, st_col = found
            series = {}
            for row in rows[header_idx + 1:]:
                if not row:
                    continue
                date = _parse_date(row[0])
                if date is None:
                    continue
                sw = row[sw_col] if sw_col < len(row) else None
                st = row[st_col] if st_col < len(row) else None
                if isinstance(sw, (int, float)) and isinstance(st, (int, float)):
                    series[date] = (float(sw), float(st))
            if series:
                _sanity_check(series)
                return series
            notes.append(f'{title}: headers found, no parseable rows')
    raise AtlantaFedError(
        'no preferred sheet parsed (sheets present: '
        + ', '.join(by_title) + ')' + ('; ' + '; '.join(notes) if notes else ''))


def fetch_switcher_premium(start='2015-01-01'):
    """Return {YYYY-MM-01: premium_pp} — switcher minus stayer wage growth.

    Only the published smoothed (3MMA) series should ever be consumed; the
    ambiguity and sanity rules above make the parser fail loudly rather than
    silently score an unsmoothed or misparsed cut.
    """
    content, _url = _workbook_bytes()
    series = parse_workbook(content)
    return {
        date: round(sw - st, 2)
        for date, (sw, st) in sorted(series.items())
        if date >= start
    }


def fetch_wgt_column(sheet_name, column_pattern, start='2015-01-01'):
    """Return {YYYY-MM-01: value} for one column of one tracker sheet.

    Used for the Occupation cut ('Professional and management' etc.), whose
    sheets are 12-month moving averages of median 12-month wage growth.
    Exactly one column may match `column_pattern` in the header row; values
    must pass the wage-growth magnitude bounds. Fails loudly on anything
    ambiguous — callers treat failure as a missing component.
    """
    import openpyxl  # deferred

    content, _url = _workbook_bytes()
    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = next((s for s in workbook.worksheets if s.title == sheet_name), None)
    if sheet is None:
        raise AtlantaFedError(f'no sheet named {sheet_name!r}')
    rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    pattern = re.compile(column_pattern, re.I)
    for row_idx, row in enumerate(rows[:10]):
        matches = [i for i, c in enumerate(row) if isinstance(c, str) and pattern.search(c)]
        if not matches:
            continue
        if len(matches) > 1:
            raise AtlantaFedError(
                f'{sheet_name}: {len(matches)} columns match {column_pattern!r}')
        col = matches[0]
        series = {}
        for row_data in rows[row_idx + 1:]:
            if not row_data:
                continue
            date = _parse_date(row_data[0])
            if date is None or date < start:
                continue
            v = row_data[col] if col < len(row_data) else None
            if isinstance(v, (int, float)):
                series[date] = float(v)
        if not series:
            raise AtlantaFedError(f'{sheet_name}: column matched but no numeric rows')
        values = sorted(series.values())
        median = values[len(values) // 2]
        if not (_MIN_PLAUSIBLE_MEDIAN <= median <= _WAGE_RANGE[1]):
            raise AtlantaFedError(
                f'{sheet_name}: magnitudes implausible (median {median})')
        return series
    raise AtlantaFedError(f'{sheet_name}: no header row matches {column_pattern!r}')
